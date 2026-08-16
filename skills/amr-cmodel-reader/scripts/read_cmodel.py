#!/usr/bin/env python3
"""Parse AMR Studio V4 .cmodel files and summarize decoded facts."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
from pathlib import Path
from typing import Any

try:
    import requests
except Exception:
    requests = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def value_of(item: Any) -> Any:
    if not isinstance(item, dict):
        return None
    for key in (
        "stringValue",
        "string_value",
        "uint32Value",
        "uint32_value",
        "int32Value",
        "int32_value",
        "doubleValue",
        "double_value",
        "floatValue",
        "float_value",
        "boolValue",
        "bool_value",
        "stringFix",
    ):
        if key in item:
            return item[key]
    combo = item.get("comboType") or item.get("combo_type")
    if isinstance(combo, dict):
        return combo.get("typeKey") or combo.get("type_key") or combo.get("typeDesc") or combo.get("type_desc")
    return None


def combo_key(item: Any) -> str | None:
    combo = item.get("comboType") or item.get("combo_type") if isinstance(item, dict) else None
    return combo.get("typeKey") or combo.get("type_key") if isinstance(combo, dict) else None


def classify_amr(main_type: str | None, sub_type: str | None = None) -> str:
    key = (main_type or "").strip()
    sub = (sub_type or "").strip()
    if key == "chassis":
        return "底盘"
    if key in {"driveWheel", "driver", "PMSMMotor", "motor"} or sub in {"diffWheel", "diffSteerWheel", "verticalSteerWheel", "horizontalSteerWheel"}:
        return "驱动单元"
    if key in {"sensor", "SENSOR", "sensorProcessor"}:
        return "传感器"
    if key in {"battery", "energyController"}:
        return "电池"
    if key in {"extendedlnterface", "extendedInterface", "ioModule"}:
        return "IO模块"
    if key in {"mainCPU", "intergratedController", "controller"}:
        return "控制器"
    if key == "screen":
        return "显示屏"
    if key == "audio":
        return "扬声器"
    if key in {"light", "lamp"}:
        return "灯带"
    if key == "button":
        return "按钮"
    if key == "actor":
        return "执行器"
    return "未分类"


def collect_components(root: dict[str, Any]) -> list[tuple[str | None, dict[str, Any]]]:
    out: list[tuple[str | None, dict[str, Any]]] = []

    def walk(group: dict[str, Any], parent: str | None = None) -> None:
        group_name = group.get("moduleGroupName") or group.get("module_group_name") or parent
        for comp in group.get("moduleComponets") or group.get("module_components") or []:
            if isinstance(comp, dict) and "$ref" not in comp:
                out.append((group_name, comp))
        for child in group.get("moreModuleInfo") or group.get("more_module_info") or []:
            if isinstance(child, dict):
                walk(child, group_name)

    for group in root.get("moreModuleInfo") or root.get("more_module_info") or []:
        if isinstance(group, dict):
            walk(group)
    return out


def extend_params(comp: dict[str, Any]) -> dict[str, Any]:
    struct = comp.get("structParam") or comp.get("struct_param") or {}
    params = struct.get("extendParams") or struct.get("extend_params") or []
    out: dict[str, Any] = {}
    for item in params if isinstance(params, list) else []:
        if isinstance(item, dict) and item.get("key"):
            out[item["key"]] = value_of(item)
    return out


def interfaces(comp: dict[str, Any]) -> list[dict[str, Any]]:
    params = comp.get("interfaceParams") or comp.get("interface_params") or {}
    groups = params.get("interfaceGroup") or params.get("interface_group") or []
    groups = [groups] if isinstance(groups, dict) else groups
    out: list[dict[str, Any]] = []
    for group in groups if isinstance(groups, list) else []:
        if not isinstance(group, dict):
            continue
        if group.get("interfaceUuid") or group.get("interface_uuid"):
            out.append(group)
        for key in ("interface", "interfaces", "arrayInterface", "array_interface"):
            nested = group.get(key)
            if isinstance(nested, list):
                out.extend(item for item in nested if isinstance(item, dict))
    return out


def flatten_base_elements(value: Any, prefix: str = "") -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if isinstance(value, dict):
        key = value.get("key")
        attr_type = value.get("type")
        if key or attr_type:
            rows.append(
                {
                    "path": prefix,
                    "key": key,
                    "type": attr_type,
                    "desc": value.get("desc"),
                    "unit": value.get("unit"),
                    "value": value_of(value),
                }
            )
        combo = value.get("comboType") or value.get("combo_type")
        if isinstance(combo, dict):
            selected = combo.get("typeKey") or combo.get("type_key")
            for group in combo.get("typeGroups") or combo.get("type_groups") or []:
                group_key = group.get("key") if isinstance(group, dict) else None
                child_prefix = f"{prefix}/{key or 'combo'}[{group_key or ''}]"
                for child_key in ("arrayCmobEle", "array_cmob_ele", "arrayAttr", "array_attr", "arrayBaseEle", "array_base_ele"):
                    children = group.get(child_key) if isinstance(group, dict) else None
                    if isinstance(children, list):
                        for child in children:
                            if selected is None or group_key == selected:
                                rows.extend(flatten_base_elements(child, child_prefix))
        for child_key in ("arrayBaseEle", "array_base_ele", "interfaceParamsArray", "interface_params_array", "linkAttrs", "link_attrs"):
            children = value.get(child_key)
            if isinstance(children, list):
                for child in children:
                    rows.extend(flatten_base_elements(child, f"{prefix}/{key or child_key}"))
    elif isinstance(value, list):
        for idx, item in enumerate(value):
            rows.extend(flatten_base_elements(item, f"{prefix}[{idx}]"))
    return rows


def private_attribute_rows(comp: dict[str, Any], module: dict[str, Any]) -> list[dict[str, Any]]:
    private = comp.get("privateAttr") or comp.get("private_attr") or {}
    groups = private.get("privateAttrs") or private.get("private_attrs") or []
    rows: list[dict[str, Any]] = []
    for group in groups if isinstance(groups, list) else []:
        if not isinstance(group, dict):
            continue
        group_key = group.get("groupKey") or group.get("group_key") or group.get("key")
        group_name = group.get("groupName") or group.get("group_name") or group.get("desc")
        elements = group.get("arrayBaseEle") or group.get("array_base_ele") or []
        for attr in flatten_base_elements(elements, str(group_key or "")):
            rows.append(
                {
                    "moduleUuid": module.get("moduleUuid"),
                    "moduleName": module.get("moduleName"),
                    "amrCategory": module.get("amrCategory"),
                    "groupKey": group_key,
                    "groupName": group_name,
                    **attr,
                }
            )
    return rows


def electrical_attribute_rows(comp: dict[str, Any], module: dict[str, Any], iface_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    ability = comp.get("interfaceAbility") or comp.get("interface_ability") or {}
    attrs = comp.get("interfaceAttrs") or comp.get("interface_attrs") or {}
    params = comp.get("interfaceParams") or comp.get("interface_params") or {}
    for iface in iface_rows:
        rows.append(
            {
                "moduleUuid": module.get("moduleUuid"),
                "moduleName": module.get("moduleName"),
                "amrCategory": module.get("amrCategory"),
                "interfaceUuid": iface.get("interfaceUuid"),
                "interfaceKey": iface.get("key"),
                "interfaceType": iface.get("type"),
                "fieldGroup": "interface",
                "path": "",
                "key": iface.get("key"),
                "type": iface.get("type"),
                "desc": iface.get("desc"),
                "value": json.dumps(iface.get("linkedInterfaceUuid") or [], ensure_ascii=False),
            }
        )
        for link_attr in iface.get("linkAttrs") or []:
            for attr in flatten_base_elements(link_attr, "linkAttrs"):
                rows.append(
                    {
                        "moduleUuid": module.get("moduleUuid"),
                        "moduleName": module.get("moduleName"),
                        "amrCategory": module.get("amrCategory"),
                        "interfaceUuid": iface.get("interfaceUuid"),
                        "interfaceKey": iface.get("key"),
                        "interfaceType": iface.get("type"),
                        "fieldGroup": "linkAttrs",
                        **attr,
                    }
                )
    for field_group, source in (("interfaceAbility", ability), ("interfaceAttrs", attrs), ("interfaceParams", params)):
        for attr in flatten_base_elements(source, field_group):
            rows.append(
                {
                    "moduleUuid": module.get("moduleUuid"),
                    "moduleName": module.get("moduleName"),
                    "amrCategory": module.get("amrCategory"),
                    "interfaceUuid": "",
                    "interfaceKey": "",
                    "interfaceType": "",
                    "fieldGroup": field_group,
                    **attr,
                }
            )
    return rows


def function_nodes(raw: dict[str, Any]) -> list[dict[str, str]]:
    nodes: list[dict[str, str]] = []

    def walk(item: dict[str, Any], prefix: str = "") -> None:
        node_type = str(item.get("type") or item.get("key") or "")
        desc = str(item.get("desc") or item.get("name") or "")
        path = f"{prefix}/{node_type}" if prefix and node_type else node_type or prefix
        nodes.append({"type": node_type, "desc": desc, "path": path})
        for child_key in ("childFunction", "child_function", "children"):
            children = item.get(child_key)
            if isinstance(children, list):
                for child in children:
                    if isinstance(child, dict):
                        walk(child, path)

    funcs = raw.get("function") or raw.get("functions") or []
    for func in funcs if isinstance(funcs, list) else []:
        if isinstance(func, dict):
            walk(func)
    return nodes


def decode_local(project_root: Path, cmodel: Path, work_dir: Path) -> dict[str, Any]:
    sys.path.insert(0, str(project_root / "src/backend"))
    from app.infrastructure.protobuf import decode_cmodel

    decode_dir = work_dir / cmodel.stem
    decode_dir.mkdir(parents=True, exist_ok=True)
    audit = decode_cmodel(str(cmodel), str(decode_dir))
    result: dict[str, Any] = {"audit": audit, "decodedDir": str(decode_dir)}
    for name, key in (("CompDesc.json", "full_json"), ("AbiSet.json", "abilities"), ("FuncDesc.json", "functions")):
        path = decode_dir / name
        if path.exists():
            result[key] = json.loads(path.read_text(encoding="utf-8"))
    if "full_json" not in result:
        raise RuntimeError("CMODEL_COMPDESC_MISSING: local decoder did not produce CompDesc.json")
    return result


def decode_api(api: str, cmodel: Path) -> dict[str, Any]:
    if requests is None:
        raise RuntimeError("requests is required for --api mode")
    with cmodel.open("rb") as handle:
        response = requests.post(
            f"{api.rstrip('/')}/models/upload",
            files={"file": (cmodel.name, handle, "application/octet-stream")},
            timeout=90,
        )
    if not response.ok:
        raise RuntimeError(f"upload failed: HTTP {response.status_code}: {response.text[:500]}")
    uploaded = response.json()
    project_id = uploaded.get("project_id")
    abilities: dict[str, Any] = {}
    functions: dict[str, Any] = {}
    if project_id:
        for endpoint, target in (("abilities", abilities), ("functions", functions)):
            r = requests.get(f"{api.rstrip('/')}/models/{project_id}/{endpoint}", timeout=60)
            if r.ok:
                target.update(r.json())
    return {
        "remoteProjectId": project_id,
        "full_json": uploaded.get("full_json") or {},
        "abilities": abilities,
        "functions": functions,
        "audit": uploaded.get("audit") or [],
    }


def summarize(cmodel: Path, decoded: dict[str, Any]) -> dict[str, Any]:
    full = decoded["full_json"]
    iface_index: dict[str, dict[str, Any]] = {}
    components = []
    mounting = []
    private_attrs = []
    electrical_attrs = []
    for group_name, comp in collect_components(full):
        general = comp.get("generalAttr") or comp.get("general_attr") or {}
        uuid = value_of(general.get("moduleUuid") or general.get("module_uuid"))
        name = value_of(general.get("moduleName") or general.get("module_name"))
        desc = value_of(general.get("moduleDesc") or general.get("module_desc"))
        main_type = combo_key(general.get("mainModuleType") or general.get("main_module_type"))
        sub_type = combo_key(general.get("subModuleType") or general.get("sub_module_type"))
        amr_category = classify_amr(main_type, sub_type)
        mount = {k: v for k, v in extend_params(comp).items() if k.startswith("locCoord") or k == "parentNodeUuid"}
        iface_rows = []
        for iface in interfaces(comp):
            iface_uuid = iface.get("interfaceUuid") or iface.get("interface_uuid")
            row = {
                "interfaceUuid": iface_uuid,
                "key": iface.get("key"),
                "type": iface.get("type"),
                "desc": iface.get("desc"),
                "linkedInterfaceUuid": iface.get("linkedInterfaceUuid") or iface.get("linked_interface_uuid") or [],
            }
            iface_rows.append(row)
            if iface_uuid:
                iface_index[iface_uuid] = {
                    "moduleUuid": uuid,
                    "moduleName": name,
                    "interfaceKey": row["key"],
                    "interfaceType": row["type"],
                }
        component_row = {
            "moduleUuid": uuid,
            "moduleName": name,
            "moduleDesc": desc,
            "groupName": group_name,
            "amrCategory": amr_category,
            "mainModuleType": main_type,
            "subModuleType": sub_type,
            "interfaceCount": len(iface_rows),
            "interfaces": iface_rows,
        }
        components.append(component_row)
        mounting.append({"moduleUuid": uuid, "moduleName": name, "amrCategory": amr_category, **mount})
        private_attrs.extend(private_attribute_rows(comp, component_row))
        electrical_attrs.extend(electrical_attribute_rows(comp, component_row, iface_rows))

    connections = []
    seen = set()
    missing = 0
    for comp in components:
        for iface in comp["interfaces"]:
            source_uuid = iface.get("interfaceUuid")
            for target_uuid in iface.get("linkedInterfaceUuid") or []:
                key = tuple(sorted([source_uuid or "", target_uuid or ""]))
                if key in seen:
                    continue
                seen.add(key)
                target = iface_index.get(target_uuid)
                if not target:
                    missing += 1
                connections.append({
                    "sourceModule": comp["moduleName"],
                    "sourceInterface": iface.get("key"),
                    "sourceType": iface.get("type"),
                    "targetModule": target.get("moduleName") if target else None,
                    "targetInterface": target.get("interfaceKey") if target else None,
                    "targetType": target.get("interfaceType") if target else None,
                    "targetInterfaceUuid": target_uuid,
                })

    abilities = decoded.get("abilities") or {}
    functions = decoded.get("functions") or {}
    return {
        "sourcePath": str(cmodel),
        "sha256": sha256(cmodel),
        "status": "success",
        "remoteProjectId": decoded.get("remoteProjectId"),
        "componentCount": len(components),
        "connectionCount": len(connections),
        "missingConnectionTargets": missing,
        "componentAbilityCount": len(abilities.get("componentAbility") or []) if isinstance(abilities, dict) else 0,
        "functionAbilityCount": len(abilities.get("functionAbility") or []) if isinstance(abilities, dict) else 0,
        "functionNodeCount": len(function_nodes(functions if isinstance(functions, dict) else {})),
        "components": components,
        "mounting": mounting,
        "privateAttributes": private_attrs,
        "electricalAttributes": electrical_attrs,
        "connections": connections,
        "functions": function_nodes(functions if isinstance(functions, dict) else {}),
        "audit": decoded.get("audit") or [],
    }


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    keys: list[str] = []
    for row in rows:
        for key in row:
            if key not in keys:
                keys.append(key)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=keys)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, sheets: dict[str, list[dict[str, Any]]]) -> str | None:
    if Workbook is None:
        return "openpyxl is not installed; xlsx output was skipped"
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name[:31])
        keys: list[str] = []
        for row in rows:
            for key in row:
                if key not in keys:
                    keys.append(key)
        if not keys:
            keys = ["message"]
            rows = [{"message": "无数据"}]
        ws.append(keys)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in rows:
            ws.append([json.dumps(row.get(key), ensure_ascii=False) if isinstance(row.get(key), (dict, list)) else row.get(key) for key in keys])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for idx, key in enumerate(keys, 1):
            width = min(max(len(str(key)) + 2, 12), 42)
            for cell in ws[get_column_letter(idx)]:
                if cell.value is not None:
                    width = min(max(width, min(len(str(cell.value)) + 2, 80)), 42)
            ws.column_dimensions[get_column_letter(idx)].width = width
    wb.save(path)
    return None


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help=".cmodel file or directory")
    parser.add_argument("--out", required=True, help="Output directory")
    parser.add_argument("--api", help="AMR Studio API base, e.g. http://host:8888/api/v1")
    parser.add_argument("--project-root", help="AMR Studio project root for local protobuf decode")
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--xlsx", action="store_true", help="Also write amr_cmodel_report.xlsx if openpyxl is installed")
    args = parser.parse_args()

    source = Path(args.input).expanduser()
    files = [source] if source.is_file() else sorted(source.rglob("*.cmodel"))
    files = files[: args.limit or None]
    out = Path(args.out).expanduser()
    out.mkdir(parents=True, exist_ok=True)

    results = []
    work_dir = out / "decoded"
    work_dir.mkdir(exist_ok=True)
    for file_path in files:
        try:
            if args.api:
                decoded = decode_api(args.api, file_path)
            elif args.project_root:
                decoded = decode_local(Path(args.project_root).expanduser(), file_path, work_dir)
            else:
                raise RuntimeError("Either --api or --project-root is required")
            results.append(summarize(file_path, decoded))
        except Exception as exc:
            results.append({"sourcePath": str(file_path), "sha256": sha256(file_path), "status": "error", "error": repr(exc)})

    (out / "summary.json").write_text(json.dumps({"results": results}, ensure_ascii=False, indent=2), encoding="utf-8")
    model_rows = [
        {k: r.get(k) for k in ("sourcePath", "sha256", "status", "remoteProjectId", "componentCount", "connectionCount", "missingConnectionTargets", "componentAbilityCount", "functionAbilityCount", "functionNodeCount", "error")}
        for r in results
    ]
    component_rows = [{"sourcePath": r["sourcePath"], **{k: v for k, v in c.items() if k != "interfaces"}} for r in results if r.get("status") == "success" for c in r["components"]]
    structure_rows = [{"sourcePath": row["sourcePath"], "amrCategory": row.get("amrCategory"), "moduleName": row.get("moduleName"), "moduleDesc": row.get("moduleDesc"), "mainModuleType": row.get("mainModuleType"), "subModuleType": row.get("subModuleType"), "moduleUuid": row.get("moduleUuid")} for row in component_rows]
    mounting_rows = [{"sourcePath": r["sourcePath"], **m} for r in results if r.get("status") == "success" for m in r["mounting"]]
    private_rows = [{"sourcePath": r["sourcePath"], **a} for r in results if r.get("status") == "success" for a in r["privateAttributes"]]
    connection_rows = [{"sourcePath": r["sourcePath"], **c} for r in results if r.get("status") == "success" for c in r["connections"]]
    electrical_rows = [{"sourcePath": r["sourcePath"], **a} for r in results if r.get("status") == "success" for a in r["electricalAttributes"]]
    function_rows = [{"sourcePath": r["sourcePath"], **f} for r in results if r.get("status") == "success" for f in r["functions"]]
    diagnostic_rows = [{"sourcePath": r["sourcePath"], "status": r.get("status"), "error": r.get("error"), "missingConnectionTargets": r.get("missingConnectionTargets")} for r in results if r.get("status") != "success" or r.get("missingConnectionTargets")]
    write_csv(out / "model_index.csv", model_rows)
    write_csv(out / "amr_structure.csv", structure_rows)
    write_csv(out / "components.csv", component_rows)
    write_csv(out / "mounting.csv", mounting_rows)
    write_csv(out / "private_attributes.csv", private_rows)
    write_csv(out / "connections.csv", connection_rows)
    write_csv(out / "electrical_attributes.csv", electrical_rows)
    write_csv(out / "functions.csv", function_rows)
    if args.xlsx:
        warning = write_xlsx(
            out / "amr_cmodel_report.xlsx",
            {
                "模型总览": model_rows,
                "AMR组织结构": structure_rows,
                "器件清单": component_rows,
                "安装位置": mounting_rows,
                "私有属性": private_rows,
                "电气连接关系": connection_rows,
                "电气属性": electrical_rows,
                "功能块描述": function_rows,
                "诊断": diagnostic_rows,
            },
        )
        if warning:
            print(f"WARNING: {warning}")
    print(out / "summary.json")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
